"""
Excel 문서 생성기
Pandas와 Openpyxl을 사용하여 스키마 명세서를 Excel로 출력
"""
from typing import Dict, List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path


class ExcelGenerator:
    """Excel 문서 생성 클래스"""

    def __init__(self):
        """초기화"""
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.pk_font = Font(bold=True, color="FF0000")
        self.fk_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(self, schema_data: Dict, output_path: str):
        """
        스키마 메타데이터를 Excel로 생성

        Args:
            schema_data: 스키마 메타데이터
            output_path: 출력 파일 경로
        """
        tables = schema_data.get('tables', {})

        # DataFrame 생성
        table_spec_df = self._create_table_spec_dataframe(tables)
        relations_df = self._create_relations_dataframe(tables)
        indexes_df = self._create_indexes_dataframe(tables)

        # Excel 파일로 저장
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            table_spec_df.to_excel(writer, sheet_name='TableSpec', index=False)
            relations_df.to_excel(writer, sheet_name='Relations', index=False)
            indexes_df.to_excel(writer, sheet_name='Indexes', index=False)

        # 스타일 적용
        self._apply_styles(output_path, tables)

    def _create_table_spec_dataframe(self, tables: Dict) -> pd.DataFrame:
        """
        테이블 명세 DataFrame 생성

        Args:
            tables: 테이블 메타데이터

        Returns:
            DataFrame
        """
        rows = []

        for table_name, table_data in tables.items():
            table_comment = table_data.get('comment', '')
            columns = table_data.get('columns', [])
            primary_keys = table_data.get('primary_keys', [])

            for column in columns:
                is_pk = 'Y' if column['name'] in primary_keys else 'N'
                is_nullable = 'Y' if column.get('nullable', True) else 'N'

                rows.append({
                    '테이블명': table_name,
                    '테이블설명': table_comment,
                    '컬럼명': column['name'],
                    '컬럼설명': column.get('comment', ''),
                    '데이터타입': column['type'],
                    'PK': is_pk,
                    'Nullable': is_nullable,
                    'Default': str(column.get('default', '')),
                    'AutoIncrement': 'Y' if column.get('autoincrement', False) else 'N'
                })

        return pd.DataFrame(rows)

    def _create_relations_dataframe(self, tables: Dict) -> pd.DataFrame:
        """
        관계 (FK) DataFrame 생성

        Args:
            tables: 테이블 메타데이터

        Returns:
            DataFrame
        """
        rows = []

        for table_name, table_data in tables.items():
            foreign_keys = table_data.get('foreign_keys', [])

            for fk in foreign_keys:
                constrained_cols = ', '.join(fk.get('constrained_columns', []))
                referred_cols = ', '.join(fk.get('referred_columns', []))

                rows.append({
                    '테이블명': table_name,
                    'FK명': fk.get('name', ''),
                    '외래키컬럼': constrained_cols,
                    '참조테이블': fk.get('referred_table', ''),
                    '참조컬럼': referred_cols,
                    '참조스키마': fk.get('referred_schema', '')
                })

        return pd.DataFrame(rows)

    def _create_indexes_dataframe(self, tables: Dict) -> pd.DataFrame:
        """
        인덱스 DataFrame 생성

        Args:
            tables: 테이블 메타데이터

        Returns:
            DataFrame
        """
        rows = []

        for table_name, table_data in tables.items():
            indexes = table_data.get('indexes', [])

            for idx in indexes:
                columns_str = ', '.join(idx.get('columns', []))
                is_unique = 'Y' if idx.get('unique', False) else 'N'

                rows.append({
                    '테이블명': table_name,
                    '인덱스명': idx.get('name', ''),
                    '컬럼': columns_str,
                    'Unique': is_unique
                })

        return pd.DataFrame(rows)

    def _apply_styles(self, file_path: str, tables: Dict):
        """
        Excel 스타일 적용

        Args:
            file_path: Excel 파일 경로
            tables: 테이블 메타데이터
        """
        wb = load_workbook(file_path)

        # TableSpec 시트 스타일링
        if 'TableSpec' in wb.sheetnames:
            ws = wb['TableSpec']
            self._style_sheet(ws)
            self._highlight_pk_columns(ws, tables)

        # Relations 시트 스타일링
        if 'Relations' in wb.sheetnames:
            ws = wb['Relations']
            self._style_sheet(ws)

        # Indexes 시트 스타일링
        if 'Indexes' in wb.sheetnames:
            ws = wb['Indexes']
            self._style_sheet(ws)

        wb.save(file_path)

    def _style_sheet(self, ws):
        """
        시트 기본 스타일 적용

        Args:
            ws: 워크시트 객체
        """
        # 헤더 스타일
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # 모든 셀에 테두리 적용 및 정렬
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')

        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _highlight_pk_columns(self, ws, tables: Dict):
        """
        PK 컬럼 강조 표시

        Args:
            ws: 워크시트 객체
            tables: 테이블 메타데이터
        """
        # PK='Y'인 행 찾기
        pk_col_idx = None
        table_col_idx = None
        column_col_idx = None

        # 헤더에서 컬럼 인덱스 찾기
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'PK':
                pk_col_idx = idx
            elif cell.value == '테이블명':
                table_col_idx = idx
            elif cell.value == '컬럼명':
                column_col_idx = idx

        if not all([pk_col_idx, table_col_idx, column_col_idx]):
            return

        # PK 행에 스타일 적용
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            pk_cell = row[pk_col_idx - 1]
            if pk_cell.value == 'Y':
                # 컬럼명 셀을 굵게
                column_name_cell = row[column_col_idx - 1]
                column_name_cell.font = self.pk_font

    def generate_multi_schema(self, schemas_data: List[Dict], output_path: str):
        """
        여러 스키마를 하나의 Excel로 생성

        Args:
            schemas_data: 스키마 메타데이터 리스트
            output_path: 출력 파일 경로
        """
        all_tables = {}

        for schema_data in schemas_data:
            schema_name = schema_data.get('schema', 'default')
            tables = schema_data.get('tables', {})

            # 스키마명 접두사 추가
            for table_name, table_data in tables.items():
                prefixed_name = f"{schema_name}.{table_name}"
                all_tables[prefixed_name] = table_data

        combined_schema = {
            'schema': 'multi',
            'tables': all_tables
        }

        self.generate(combined_schema, output_path)
